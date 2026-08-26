import datetime
import json
import logging
import os
from concurrent.futures import ThreadPoolExecutor
from logging.handlers import RotatingFileHandler

from flask import Flask, jsonify, render_template, request, send_from_directory
from openpyxl import Workbook, load_workbook
from werkzeug.exceptions import HTTPException

from scraper.pjud_client import COMPETENCIAS, LIBRO_TIPO, PjudSession

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
RESUMEN_XLSX = os.path.join(OUTPUT_DIR, "resumen.xlsx")
LOG_DIR = os.path.join(BASE_DIR, "logs")
LOG_FILE = os.path.join(LOG_DIR, "app.log")

os.makedirs(LOG_DIR, exist_ok=True)

logger = logging.getLogger("pjud")
logger.setLevel(logging.DEBUG)
_formato = logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s")
_file_handler = RotatingFileHandler(LOG_FILE, maxBytes=2_000_000, backupCount=5, encoding="utf-8")
_file_handler.setFormatter(_formato)
_console_handler = logging.StreamHandler()
_console_handler.setFormatter(_formato)
logger.addHandler(_file_handler)
logger.addHandler(_console_handler)

app = Flask(__name__)


@app.errorhandler(Exception)
def manejar_error_no_controlado(err):
    """Red de seguridad: cualquier excepcion no atrapada queda en logs/app.log y
    se devuelve como JSON (nunca la pagina HTML de error por defecto de Flask),
    para que el formulario pueda mostrar el mensaje en vez de fallar al parsear
    la respuesta como JSON."""
    if isinstance(err, HTTPException):
        return err
    logger.exception("Error no controlado en %s %s", request.method, request.path)
    return jsonify({"error": f"Error interno ({err.__class__.__name__}). Revisa logs/app.log para el detalle."}), 500

# Playwright (via la API sincrona) es afin a un unico hilo del sistema operativo:
# el objeto no puede usarse desde un hilo distinto al que lo creo. Como el
# servidor de desarrollo de Flask puede atender cada request en un hilo nuevo,
# toda interaccion con la pagina se despacha a este unico hilo dedicado.
_executor = ThreadPoolExecutor(max_workers=1)
_session: PjudSession | None = None


def _get_session_sync() -> PjudSession:
    global _session
    if _session is None:
        _session = PjudSession(headless=False)
    return _session


def run_on_session(fn):
    def task():
        return fn(_get_session_sync())

    return _executor.submit(task).result()


@app.route("/")
def index():
    try:
        cortes = run_on_session(lambda s: s.get_cortes())
    except Exception:
        logger.exception("Error al iniciar la sesion de PJUD / obtener cortes")
        raise
    return render_template(
        "index.html",
        competencias=list(COMPETENCIAS.keys()),
        libro_tipo=LIBRO_TIPO,
        cortes=cortes,
    )


@app.route("/tribunales")
def tribunales():
    competencia = request.args.get("competencia", "")
    corte = request.args.get("corte", "")
    if competencia not in COMPETENCIAS or not corte:
        return jsonify([])
    try:
        resultado = run_on_session(lambda s: s.get_tribunales(competencia, corte))
        return jsonify(resultado)
    except Exception:
        logger.exception("Error al obtener tribunales (competencia=%s, corte=%s)", competencia, corte)
        return jsonify({"error": "No se pudo obtener el listado de tribunales. Revisa logs/app.log."}), 500


@app.route("/consultar", methods=["POST"])
def consultar():
    payload = request.get_json(force=True)
    competencia = payload.get("competencia")
    corte = payload.get("corte")
    tribunal = payload.get("tribunal")
    libro_tipo = payload.get("libro_tipo")
    rol = payload.get("rol")
    anio = payload.get("anio")

    if competencia not in COMPETENCIAS or not all([corte, tribunal, libro_tipo, rol, anio]):
        return jsonify({"error": "Faltan datos de busqueda"}), 400

    logger.info(
        "Consultando causa: competencia=%s corte=%s tribunal=%s libro_tipo=%s rol=%s anio=%s",
        competencia, corte, tribunal, libro_tipo, rol, anio,
    )

    carpeta_documentos = os.path.join(
        OUTPUT_DIR, competencia, f"{libro_tipo}-{rol}-{anio}", "documentos"
    )

    try:
        resultado = run_on_session(
            lambda s: s.buscar_causa(
                competencia, corte, tribunal, libro_tipo, rol, anio,
                carpeta_documentos=carpeta_documentos,
            )
        )
    except Exception:
        logger.exception(
            "Error al buscar la causa %s-%s-%s (competencia=%s, tribunal=%s)",
            libro_tipo, rol, anio, competencia, tribunal,
        )
        return jsonify({"error": "Error al consultar la causa en PJUD. Revisa logs/app.log para el detalle."}), 500

    # Las rutas de archivo que devuelve el scraper son absolutas (necesarias para
    # guardar en disco); para exponerlas al formulario se convierten a rutas
    # relativas a OUTPUT_DIR, que es lo que sirve la ruta /descargas/<ruta>.
    for doc in resultado.get("documentos", []):
        if doc.get("archivo"):
            doc["url"] = "/descargas/" + os.path.relpath(doc["archivo"], OUTPUT_DIR).replace("\\", "/")
            doc["archivo"] = os.path.relpath(doc["archivo"], BASE_DIR)

    try:
        ruta_json = guardar_resultado(competencia, libro_tipo, rol, anio, resultado)
        resultado["guardado_en"] = ruta_json
    except Exception:
        logger.exception("Error al guardar el resultado de %s-%s-%s", libro_tipo, rol, anio)
        resultado["guardado_en"] = None
        resultado["aviso"] = "La consulta se realizo pero no se pudo guardar en output/. Revisa logs/app.log."

    return jsonify(resultado)


@app.route("/descargas/<path:ruta>")
def descargas(ruta):
    return send_from_directory(OUTPUT_DIR, ruta)


def guardar_resultado(competencia, libro_tipo, rol, anio, resultado):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    carpeta = os.path.join(OUTPUT_DIR, competencia, f"{libro_tipo}-{rol}-{anio}")
    os.makedirs(carpeta, exist_ok=True)
    ruta_json = os.path.join(carpeta, "datos.json")
    with open(ruta_json, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    if os.path.exists(RESUMEN_XLSX):
        wb = load_workbook(RESUMEN_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Consultas"
        ws.append(["Fecha consulta", "Competencia", "Libro/Tipo", "Rol", "Año", "Encontrada", "Resumen"])

    resumen_txt = " | ".join(resultado.get("resumen", [])) if resultado.get("encontrada") else ""
    ws.append(
        [
            datetime.datetime.now().isoformat(timespec="seconds"),
            competencia,
            libro_tipo,
            rol,
            anio,
            resultado.get("encontrada", False),
            resumen_txt,
        ]
    )
    wb.save(RESUMEN_XLSX)
    return os.path.relpath(ruta_json, BASE_DIR)


if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    app.run(debug=False, port=5000)
